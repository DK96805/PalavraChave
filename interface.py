import os
import PyPDF2
from docx import Document
from openpyxl import load_workbook
import tkinter as tk
from tkinter import filedialog
from tkinter import scrolledtext


def search_word_in_pdf(file_path, search_word, result_text):
    with open(file_path, 'rb') as file:
        pdf = PyPDF2.PdfReader(file)
        for page_number, page in enumerate(pdf.pages, start=1):
            text = page.extract_text()
            lines = text.split('\n')
            for line_number, line in enumerate(lines, start=1):
                if search_word in line.strip():
                    result_text.insert(tk.END, f'Palavra encontrada no arquivo PDF: {file_path}\n')
                    result_text.insert(tk.END, f'Página: {page_number}\n')
                    result_text.insert(tk.END, f'Linha: {line_number}\n')
                    result_text.insert(tk.END, '---\n')

def search_word_in_word(file_path, search_word, result_text):
    filename = os.path.basename(file_path)
    if filename.startswith('~$'):
        return

    doc = Document(file_path)
    for paragraph_number, paragraph in enumerate(doc.paragraphs, start=1):
        if search_word in paragraph.text:
            result_text.insert(tk.END, f'Palavra encontrada no arquivo Word: {file_path}\n')
            result_text.insert(tk.END, f'Página: {paragraph_number}\n')
            result_text.insert(tk.END, '---\n')

def search_word_in_excel(file_path, search_word, result_text):
    workbook = load_workbook(file_path)
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            for column_number, cell_value in enumerate(row, start=1):
                if isinstance(cell_value, str) and search_word in cell_value:
                    result_text.insert(tk.END, f'Palavra encontrada no arquivo Excel: {file_path}\n')
                    result_text.insert(tk.END, f'Planilha: {sheet}\n')
                    result_text.insert(tk.END, f'Linha: {row_number}\n')
                    result_text.insert(tk.END, f'Coluna: {column_number}\n')
                    result_text.insert(tk.END, '---\n')

def show_notification():
    notification = tk.Toplevel()
    notification.title("Notificação")
    notification.geometry("300x100")
    notification_label = tk.Label(notification, text="Procurando palavras-chave...")
    notification_label.pack(pady=30)
    # Defina a duração desejada para a notificação em milissegundos (por exemplo, 2000 para 2 segundos)
    duration = 2000
    notification.after(duration, notification.destroy)


def search_word_in_directory(directory_path, search_word, result_text):
    no_match_found = True  # Inicialmente, consideramos que não há correspondências

    for root, dirs, files in os.walk(directory_path):
        for file in files:
            file_path = os.path.join(root, file)
            if file.endswith('.pdf'):
                search_word_in_pdf(file_path, search_word, result_text)
            elif file.endswith('.docx'):
                search_word_in_word(file_path, search_word, result_text)
            elif file.endswith('.xlsx'):
                search_word_in_excel(file_path, search_word, result_text)

            # Verifica se foram encontradas correspondências
            if result_text.get("1.0", tk.END).strip():
                no_match_found = False  # Correspondência encontrada

    # Verifica se nenhuma correspondência foi encontrada
    if no_match_found:
        no_match_window = tk.Toplevel()
        no_match_window.title("Aviso")
        no_match_label = tk.Label(no_match_window, text="Nenhuma correspondência encontrada.")
        no_match_label.pack(pady=20)



def select_directory(input_entry, result_text):
    directory_path = filedialog.askdirectory()
    if directory_path:
        search_word = input_entry.get()
        show_notification()
        result_text.delete('1.0', tk.END)  # Limpar o texto de resultados
        search_word_in_directory(directory_path, search_word, result_text)

def create_window():
    window = tk.Tk()
    window.title("Palavra-CHAVE")
    window.iconbitmap('imagens/cranio-e-ossos.ico')  # Substitua pelo caminho para o seu arquivo .ico

    title_label = tk.Label(window, text="Programa criado por DK96805. ")
    title_label.pack(pady=10)

    title_label = tk.Label(window, text="*ATENÇÃO*, a busca deve respeitar acentos, letras MAIÚSCULAS e MINUSCULAS. ")
    title_label.pack(pady=11)

    input_frame = tk.Frame(window)
    input_frame.pack(pady=10)

    input_label = tk.Label(input_frame, text="Palavra-chave:")
    input_label.pack(side=tk.LEFT)

    input_entry = tk.Entry(input_frame, width=30)
    input_entry.pack(side=tk.LEFT)

    select_button = tk.Button(window, text="Selecionar Diretório", command=lambda: select_directory(input_entry, result_text))
    select_button.pack(padx=10, pady=10)

    result_frame = tk.Frame(window)
    result_frame.pack(pady=10)

    result_label = tk.Label(result_frame, text="Resultados:")
    result_label.pack()

    result_text = scrolledtext.ScrolledText(result_frame, width=60, height=15)
    result_text.pack()

    global no_match_found
    no_match_found = tk.BooleanVar(value=False)

    window.mainloop()

create_window()
