import os
import PyPDF2
from docx import Document
from openpyxl import load_workbook


def search_word_in_pdf(file_path, search_word):
    with open(file_path, 'rb') as file:
        pdf = PyPDF2.PdfReader(file)
        for page_number, page in enumerate(pdf.pages, start=1):
            text = page.extract_text()
            lines = text.split('\n')
            for line_number, line in enumerate(lines, start=1):
                if search_word in line.strip():
                    print(f'Palavra encontrada no arquivo PDF: {file_path}')
                    print(f'Página: {page_number}')
                    print(f'Linha: {line_number}')
                    print('---')
    return

def search_word_in_word(file_path, search_word):
    filename = os.path.basename(file_path)
    if filename.startswith('~$'):
        return

    doc = Document(file_path)
    for paragraph_number, paragraph in enumerate(doc.paragraphs, start=1):
        if search_word in paragraph.text:
            print(f'Palavra encontrada no arquivo Word: {file_path}')
            print(f'Página: {paragraph_number}')
            print('---')
    return

def search_word_in_excel(file_path, search_word):
    workbook = load_workbook(file_path)
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            for column_number, cell_value in enumerate(row, start=1):
                if isinstance(cell_value, str) and search_word in cell_value:
                    print(f'Palavra encontrada no arquivo Excel: {file_path}')
                    print(f'Planilha: {sheet}')
                    print(f'Linha: {row_number}')
                    print(f'Coluna: {column_number}')
                    print('---')
    return

def search_word_in_directory():
    directory_path = input('Informe o diretório onde se encontram os arquivos: ')
    search_word = input('Informe a palavra-chave buscada: ')
    print('---')

    for root, dirs, files in os.walk(directory_path):
        for file in files:
            file_path = os.path.join(root, file)
            if file.endswith('.pdf'):
                search_word_in_pdf(file_path, search_word)
            elif file.endswith('.docx'):
                search_word_in_word(file_path, search_word)
            elif file.endswith('.xlsx'):
                search_word_in_excel(file_path, search_word)


search_word_in_directory()
